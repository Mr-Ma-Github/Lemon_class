#-*-coding:utf-8-*-
#@作者：haiyu.ma
#@创建日期：2019-12-06 18:00 
from GC_API.tools import project_path
from openpyxl import load_workbook
import pandas as pd
from GC_API.tools.case_config import ReadConfig

# wb=load_workbook(project_path.test_case_path)#Open the given filename and return the workbook
# sheet=wb['python']#传表单名 定位表单   返回一个表单对象
# res=sheet.cell(1,1).value #定位单元格   根据行列值
class GetData:
    Cookie=None
    Authorization = None
    # NoRegTel=pd.read_excel(project_path.test_case_path,sheet_name='init').ix[0,0]
    # 参照init表单查看那我们这个变量的用处
    loan_id=None
    check_list=eval(ReadConfig.get_config(project_path.case_config_path,'CHECKLEAVEAMOUNT','check_list'))
    NoRegTel = load_workbook(project_path.test_data_path)['init'].cell(2,1).value
    normal_tel = load_workbook(project_path.test_data_path)['init'].cell(3, 1).value
    admin_tel = load_workbook(project_path.test_data_path)['init'].cell(4, 1).value
    loan_member_id = load_workbook(project_path.test_data_path)['init'].cell(5, 1).value
    member_Id = load_workbook(project_path.test_data_path)['init'].cell(6, 1).value
    pwd=123456
    # print(memberId)
# setattr(GetCookie,'Cookie','123456')#set attribute 设置属性值
# print(hasattr(GetCookie,'Cookie'))#has attribute  判断属性值是否存在
# delattr(GetCookie,'Cookie')# delete attribute  删除这个属性
# print(getattr(GetCookie,'Cookie'))#get attribute 获取属性值

# print(getattr(GetData,'NoRegTel'))
# df=pd.read_excel(project_path.test_case_path,sheet_name='init')
# print(df.ix[0,0])
# print(pd.read_excel(project_path.test_case_path,sheet_name='init').ix[0,0])
# print(GetData.wb)
