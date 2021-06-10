#-*-coding:utf-8-*-
from openpyxl import load_workbook
from class_1120.read_config import ReadConfig


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):  # ,mode="all"
        """mode:控制是否执行所有的用例  默认值为all 如果为all就执行所有用例，否则进入分值判断
           mode只是一个参数名，mode的值只能输入all  列表  """

        # 从配置文件中读取mode值
        mode = ReadConfig().read_config(R'C:\Users\haiyu.ma\PycharmProjects\lemon_class\APIAutomation\class_1120\case.config', 'MODE', 'mode')

        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for item in range(1, sheet.max_row+1):
            sub_data = {}
            sub_data["case_id"] = sheet.cell(item, 1).value
            # url=sheet.cell(item,2).value
            sub_data["url"] = sheet.cell(item,3).value
            # data.txt=sheet.cell(item,3).value
            sub_data["data.txt"] = sheet.cell(item, 4).value
            # excepted=sheet.cell(item,4).value
            sub_data["excepted"] = sheet.cell(item, 5).value
            # method=sheet.cell(item,1).value
            sub_data["method"] = sheet.cell(item, 2).value
            # print(url,data.txt,excepted,method)
            test_data.append(sub_data)

        # 根据mode值去进行判断
        if mode == "all": # 执行所有的用例
            final_data = test_data
        else:  # [1,2,3,4]
            final_data = []
            for item in test_data:  # 遍历所有的测试数据
                if item["case_id"] in eval(mode):  # 文件中用例的序号
                    final_data.append(item)
        return final_data  # 返回获取到的数据

if __name__ == '__main__':
    print(DoExcel("Excel_study.xlsx", "python").get_data())


