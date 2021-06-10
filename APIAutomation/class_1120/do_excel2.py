#-*-coding:utf-8-*-
from openpyxl import load_workbook


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]
        self.max_row = self.sheet_obj.max_row
        #获取一个表单对象

    def get_data(self, i, j):
        '''根据传入的坐标来获取值'''
        return self.sheet_obj.cell(i, j).value

    # def get_max_row(self):
    #     '''根据表单对象获取最大的行数'''
    #     return self.sheet_obj.max_row


if __name__ == '__main__':
    res = DoExcel("Excel_study.xlsx", "python")
    # print(res.get_data(1, 1))
    print(res.get_data(1, 1), res.get_data(1, 2), str(res.get_data(1, 3)), eval(res.get_data(1, 4)))
    list_data = []
    dict_data = {}
    dict_data["id"] = res.get_data(1, 1)
    dict_data["mothod"] = res.get_data(1, 2)
    dict_data["url"] = res.get_data(1, 3)
    dict_data["data"] = res.get_data(1, 4)
    list_data.append(dict_data)
    print(list_data)