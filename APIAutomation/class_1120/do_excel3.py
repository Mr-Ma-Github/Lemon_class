#-*-coding:utf-8-*-
#!/usr/bin/python
#-*- coding:utf-8 -*-
#@Author：zhuxiujie

from openpyxl import load_workbook

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open_excel(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        wb.close()
        return sheet

    def get_header(self):
        """获取第一行的标题行"""
        sheet = self.open_excel()
        header = []     #存储标题行
        for j in range(1, sheet.max_column+1):  # 1,2,3,4,5
            header.append(sheet.cell(1, j).value)
        return header

    def get_data(self):
        """根据嵌套循环读取数据"""
        sheet = self.open_excel()
        header = self.get_header()  # 拿到header  是一个列表  索引是从0开始的
        test_data = []
        for i in range(2, sheet.max_row+1):  # 2,3,4,5,6
            sub_data = {}
            for j in range(1, sheet.max_column+1):  # 1,2,3,4,5
                sub_data[header[j-1]] = sheet.cell(i, j).value  # j-1是因为header是列表，从下标0开始取值
            test_data.append(sub_data)
        return test_data

    def write_back(self, row, col, result):  # 写回数据
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = result
        wb.save(self.file_name)

if __name__ == '__main__':
    # print(DoExcel('python.xlsx', 'python').get_header())
    print(DoExcel('python.xlsx', 'python').get_data())
    # DoExcel('python.xlsx', 'python').write_back(1, 6, '测试1')
    # DoExcel('python.xlsx', 'python').write_back(2, 6, '测试2')
