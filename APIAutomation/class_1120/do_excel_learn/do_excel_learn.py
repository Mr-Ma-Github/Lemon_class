# -*-coding:utf-8-*-
# 手工操作Excel操作流程：
# 1.打开Excel文件（路径+文件名）
# 2.定位表单
# 3.使用行、列去定位要读取的数据
# 4.关闭文件

# python操作Excel，工具：
# -openpyxl：仅支持 xlsx 格式的 读写
# -tablib：支持多种格式的 读写。xlsx，xls，csv，json，yaml，html，pandas
# -xlrd：经典的Excel 读取 库
# -pandas：功能强，太臃肿了

# 安装：pip install openpyxl

from openpyxl import load_workbook
import os



directory_path = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(directory_path, "Excel_study.xlsx")
# print(file_path)

# 1.打开Excel文件：读取之前一定要关闭该文件
# wb = load_workbook(file_path)  # Open the given filename and return the workbook
# print(wb)
# 不直接去获取_sheets 属性，称为私有属性
# print(wb._sheets)
# sheetnames 和 _sheets 有什么区别？？
# print(wb.sheetnames)
# sheetnames 列表当中当中存储的是字符串。_sheets 列表里面存储的是对象
# 获取被激活的表单：active 是表示被激活，被选择的sheet
# active_sheet = wb.active
# print(active_sheet)
# 获取所有表单的正确用法：
# print(wb.worksheets)
# 获取某一个表单：
# 1)通过索引获取
# sheet = wb.worksheets[0]
# print(sheet)
# 2)通过表单名称获取
# sheet2 = wb.get_sheet_by_name("python")  # 过时了
# print(sheet2)
# 3.通过名字获取表单用法：--建议使用这个
# sheet3 = wb['python']
# pycharm 不支持 sheet.属性的提示

# 2.定位表单
# sheet = wb['python']  # 传表单名    返回一个表单对象
# sheet = wb.worksheets[0]  # 注意使用这种方法pycharm会有智能提示

# 3.定位单元格   根据行列值：行列值从1开始
# res = sheet.cell(1, 2)  # 获取的是对象
# print(res)
# res1 = sheet.cell(1, 2).value  # (行，列) # 获取的是值
# print(res1)
# print("最大行数:{}".format(sheet.max_row))
# print("最大列数:{}".format(sheet.max_column))

# 获取表单里面的某一行：
# print(sheet[1])
# # # 获取值：
# for row in sheet[1]:
#     print(row.value)

# 获取表单里面的某一列：
# print(sheet['A'])
# 获取多行：1到3行，第三行是包含的
# print(sheet[1:3])
# 获取所有的数据：
# total_data = list(sheet.rows)   # list(sheet.rows)[1:3]  支持切片
# print(total_data)
# for row in total_data:
#     for cell in row:
#         print(cell.value)

# 写入一个单元格:
# sheet.cell(1, 6).value = "写入数据"

# 写入后需要保存:
# wb.save(file_path)

# 关闭文件
# wb.close()

# 数据从Excel中拿出来是什么类型？
# 数字依然是int、folat，其他都是字符串
# print("method:{},类型是{}".format(sheet.cell(1, 2).value, type(sheet.cell(1, 2).value)))
# print("url:{},类型是{}".format(sheet.cell(1, 3).value, type(sheet.cell(1, 3).value)))
# print("data:{},类型是{}".format(sheet.cell(1, 4).value, type(sheet.cell(1, 4).value)))
# print("excepted:{},类型是{}".format(sheet.cell(1, 5).value, type(sheet.cell(1, 5).value)))


# eval() 把数据类型转换成 原本数据类型
# s = "True"
# print(eval(s), type(eval(s)))
# b = '{"name": "推荐食谱", "1": "症状", "name1": "浑身忽冷忽热"}'
# print(b, type(b))
# print(eval(b), type(eval(b)))

# zip()
a = [1,2,3]
b = [4,5,6]
# zipped = dict(zip(a, b))
# print(zipped)
body_data = []
for i in a:
    body_data.append(i)
    data_dict = dict(zip(b, body_data))
    print(data_dict)

# 读取数据  方式一 方式二  以后肯定是方法一
# 方法一：一次性读取所有的数据，对内存的要求比较高
# 方法二：需要用的时候读取所有的数据，对磁盘读写要求比较高
# 方法三：比较复杂一点
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

class DoExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def open_file(self) -> Worksheet:  # 加上返回值类型，会有智能提示
        # 在函数或者方法的后面加 -> 类型，表示此函数返回值的类型
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        return sheet

    def get_header(self):
        sheet = self.open_file()
        header = []
        # 方法一：
        for column in sheet[1]:
            header.append(column.value)
        return header
        # 方法二：
        # for column in range(1, sheet.max_column+1):
        #     # print(sheet.cell(1, column).value)
        #     header.append(sheet.cell(1, column).value)
        # return header

    def get_data(self):
        sheet = self.open_file()
        header = self.get_header()
        test_data = []
        # 方法一：
        rows = list(sheet.rows)[1:]  # 与列表切片有区别，是取尾的
        for row in rows:
            data = []
            for cell in row:
                data.append(cell.value)
                # 两个列表结合成字典，要用zip
                global data_dict
                data_dict = dict(zip(header,data))
            test_data.append(data_dict)
        return test_data
        # 方法二：
        # for row in range(2, sheet.max_row + 1):
        #     data = {}
        #     for column in range(1, sheet.max_column + 1):
        #         data[header[column - 1]] = sheet.cell(row, column).value
        #     test_data.append(data)
        # return test_data

    @staticmethod
    def write_workbook(file, sheetname, row, column, data):
        wb = load_workbook(file)
        sheet = wb[sheetname]
        sheet.cell(row, column).value = data
        wb.save(file)
        wb.close()

# if __name__ == '__main__':
    # import os
    # path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # file_path = os.path.join(path,"python.xlsx")
    # # print(DoExcel(file_path, "python").get_header())
    # DoExcel(file_path, "python").write_workbook(file_path, "python", 4, 6, "new value")
    # print(DoExcel(file_path, "python").get_data())


