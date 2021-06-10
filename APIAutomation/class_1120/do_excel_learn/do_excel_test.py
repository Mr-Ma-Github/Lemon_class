# -*-coding:utf-8-*-
# @作者：haiyu.ma
# @创建日期：2021-04-30 16:06 
# @Software：PyCharm
# ----------------------------------------------------------------------------
from openpyxl import load_workbook

class DoExcel:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def get_sheet(self):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        return sheet

    def read_header(self):
        sheet = self.get_sheet()
        header_data = []
        # 方式一：
        # for cell in sheet[1]:
        #     header_data.append(cell.value)
        # print(header_data)
        # 方式二：
        for column in range(1, sheet.max_column+1):
            header_data.append(sheet.cell(1, column).value)
        return header_data

    def read_body(self):
        sheet = self.get_sheet()
        test_data = []
        header = self.read_header()
        # 方式一：
        # for rows in list(sheet.rows)[1:]:
        #     body_data = []
        #     for cell in rows:
        #         body_data.append(cell.value)
        #     data_dict = dict(zip(header, body_data))
        #     test_data.append(data_dict)
        # print(test_data)
        # 方式二：
        test_data = []
        for row in range(2, sheet.max_row + 1):  # 2,3,4,5
            data = {}
            for column in range(1, sheet.max_column + 1):  # 1,2,3,4,5,6
                data[header[column - 1]] = sheet.cell(row, column).value
            test_data.append(data)
        print(test_data)



if __name__ == '__main__':
    filename = r'C:\Users\haiyu.ma\PycharmProjects\lemon_class\APIAutomation2021\class_1120\do_excel_learn\Excel_study.xlsx'
    # DoExcel(filename,'python').read_header()
    DoExcel(filename, 'python').read_body()
