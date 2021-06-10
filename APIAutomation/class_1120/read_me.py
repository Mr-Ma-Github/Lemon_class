#-*-coding:utf-8-*-
# exxel处理测试数据 测试结果 pip install openpyxl
# ddt   pip install ddt
test_data=[{"url":"http://119.23.241.154:8080/futureloan/mvc/api/member/login",
"data.txt": {"mobilephone": "18688773467", "pwd": "123456"}, "excepted": "10001", "method": "get"},
{"url": "http://119.23.241.154:8080/futureloan/mvc/api/member/login",
"data.txt": {"mobilephone": "18688773467", "pwd": "325"}, "excepted": "20111", "method": "post"},
{"url": "http://119.23.241.154:8080/futureloan/mvc/api/member/recharge",
"data.txt": {"mobilephone": "18688773467", "amount": "1000"}, "excepted": "10001", "method": "get"},
{"url": "http://119.23.241.154:8080/futureloan/mvc/api/member/recharge",
"data.txt": {"mobilephone": "18688773467", "amount": "-100"}, "excepted": "20117", "method": "get"}]

# for item in test_data:
#     print(item["url"])
#

# list=[{1:"fsf",2:"gfdg"},{3:"ter",4:"bvvcn"}]
# print([i[j] for i in list for j in i])
# for i in list:
#     for j in i:
#         print(i[j])

list=[{1:"fsf",2:"gfdg"},{1:"ter",2:"bvvcn"}]
for i in list:
    print(i[1])


# dict={1:"fsf","gd":"gfdg"}
# for i in dict:
#     print(i)

# list=[{1:"fsf",2:"gfdg"},{3:"ter",4:"bvvcn"}]
# print(list[0][1])

# 取列表中字典的值
# list = [{"name": "推荐食谱", "1": "症状", "name1": "浑身忽冷忽热"},
#         {"name": "绿豆薏米饭"}, {"name": "芝麻"}]
# res = [item[key] for item in list for key in item]
# print(res)
# for item in list:
#     for key in item:
#         print(item[key])
from openpyxl import load_workbook
wb=load_workbook("Excel_study.xlsx")#Open the given filename and return the workbook
sheet=wb['python']
method=sheet.cell(1,1).value
url=sheet.cell(1,2).value
print(method,url)

