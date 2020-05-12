#-*-coding:utf-8-*-
#@作者：haiyu.ma
#@创建日期：2019-12-03 18:10
from openpyxl import load_workbook
from API_AUTO.tools.case_config import ReadConfig
from API_AUTO.tools.project_path import *
from API_AUTO.tools.get_data import GetData
from API_AUTO.tools.do_Regular import DoRegular
class DoExcel:

    @classmethod   #可以设置成类函数
    def do_excel(cls,file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig.get_config(case_config_path, 'MODE', 'mode'))
        NoRegTel=getattr(GetData,"NoRegTel")#从GetData里拿到的数据  #这里也可以直接拿Excel里的数据
        test_data=[]
        for key in mode.keys():  # 遍历这个存在配置文件里面的字典
            sheet=wb[key]#表单名
            # print(mode[key])
            if mode[key]=="all":
                for item in range(2,sheet.max_row+1):
                    # print(item)
                    sub_data={}
                    sub_data["case_id"] = sheet.cell(item,1).value
                    sub_data["url"] = sheet.cell(item,2).value
                    # NoRegTel
                    if sheet.cell(item,3).value.find("${NoRegTel}")!=-1:#有找到这个 ${tel_1}
                        sub_data["data"]=sheet.cell(item,3).value.replace(("${NoRegTel}"),str(NoRegTel))
                        NoRegTel=NoRegTel+1
                    else:#如果没有找到的话
                        sub_data["data"] =DoRegular.do_regular(str(sheet.cell(item, 3).value))
                    # 这里是sql语句的处理
                    if sheet.cell(item,4).value != None:
                        if sheet.cell(item, 4).value.find("${normal_tel}") != -1:
                            sub_data["check_sql"] = DoRegular.do_regular(str(sheet.cell(item, 4).value))
                    else:
                        sub_data["check_sql"]=sheet.cell(item,4).value
                    # sub_data["check_sql"] = DoRegular.do_regular(str(sheet.cell(item, 4).value))#这样会有问题，如果处理None和str
                    sub_data["title"]= sheet.cell(item,5).value
                    sub_data["http_method"]= sheet.cell(item,6).value
                    sub_data["expected"] = sheet.cell(item,7).value
                    sub_data["sheet_name"]=key
                    test_data.append(sub_data)
                    cls.update_tel(NoRegTel,file_name,'init')#更新手机号
            else:
                for case_id in mode[key]:
                    print(case_id)
                    sub_data = {}
                    sub_data["case_id"] = sheet.cell(case_id+1, 1).value
                    sub_data["url"] = sheet.cell(case_id+1, 2).value
                    if sheet.cell(case_id+1,3).value.find("${NoRegTel}")!=-1:#有找到这个 ${tel_1}
                        sub_data["data"]=sheet.cell(case_id+1,3).value.replace(("${NoRegTel}"),str(NoRegTel))
                        NoRegTel=NoRegTel+1
                    else:#如果没有找到的话
                        sub_data["data"] = DoRegular.do_regular(str(sheet.cell(case_id+1, 3).value))
                    #sql语句的处理
                    # 因为有的时候Excel里面没有任何数据，所以要先判断决定是否要用正则处理数据，如果是None正则无法处理
                    # 正则无法处理，另行写代码处理
                    if sheet.cell(case_id+1,4).value != None:
                        if sheet.cell(case_id+1, 4).value.find("${normal_tel}") != -1:
                            sub_data['check_sql']=DoRegular.do_regular(str(sheet.cell(case_id+1,4).value))
                    else:
                        sub_data['check_sql']=sheet.cell(case_id+1,4).value
                    # sub_data['check_sql']=DoRegular.do_regular(str(sheet.cell(case_id+1,4).value))
                    sub_data["title"] = sheet.cell(case_id+1, 5).value
                    sub_data["http_method"] = sheet.cell(case_id+1, 6).value
                    sub_data["expected"] = sheet.cell(case_id+1, 7).value#添加了一个期望值到测试数据里面
                    sub_data["sheet_name"] = key
                    test_data.append(sub_data)
                    cls.update_tel(NoRegTel + 2, file_name, 'init')  # 更新手机号
        return(test_data)
    @staticmethod
    def write_back(file_name,sheet_name,row,col,result):#写回数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(row,col).value=result
        sheet.cell(row,col).value = result
        wb.save(file_name)
    @classmethod
    def update_tel(cls,tel,file_name,sheet_name):
        '''更新Excel里面的手机号数据'''
        wb = load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(2,1).value=tel
        wb.save(file_name)



if __name__ == '__main__':
    # web=DoExcel().do_excel(R"C:\Users\haiyu.ma\PycharmProjects\lemon_class\API_AUTO\test_data\test_data.xlsx","login")
    web2=DoExcel.do_excel(test_case_path)
    #mode={"login":"all","register":[1],"recharge":[1,2]}
    for i in web2:
        if i['check_sql']!=None:
            print(eval(i['check_sql'])['sql'])
    # web3 = DoExcel().write_back(test_case_path,'login',3,'jfksl','kfh')
    #
    # web2=DoExcel.do_excel(test_case_path)
    print(web2)
    # print(len(web2))